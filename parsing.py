##------------------------------------------PARSER----------------------------------------------------------------
import json
import triage as t
import openpyxl as op
import os



#-----------------------------------------------------------------------------------------------------------------
#---------------------Lecture fichier Json + création dico id pour les sommets------------------------------------
#-----------------------------------------------------------------------------------------------------------------

#### preque obsolete

def Json_parsing(Json_file) :

    rep = []
    no_tools = 1
    d_id_tools = {}
    d_process_tools = {}
    id = 0
    col = " "

    #ouverture du fichier Json 
    with open(Json_file) as file :
        data = json.load(file)

    #on met dans rep les process et l'outils associé l'un a coté de l'autre
    for process in data :
        rep.append([process])## me donne le nom des process
        rep.append(data[process]['tools'])## me donne les outils

    #on remplace les cases [] signifiant qu'il n'y a pas d'outils par "no_tools" et le numéro d'apparition
    for i in range(1, len(rep),1) :
        if len(rep[i]) == 0 :
            
            rep[i] = ["no_tools_" + str(no_tools)]
            no_tools += 1

    #on forme les deux dico :
    #le premier avec les process et l'outil associé
    #le deuxième avec l'id des outils en fonction de leur ordre d'apparition dans le fichier
    for i in range(0,len(rep),2) :

        exist = False
        
        #ici le process i va prendre l'outils i+1 en valeur
        #o rappelle que dans rep le process et l'outil associé se suivent dans la liste

        d_process_tools[col.join(rep[i])] = col.join(rep[i+1])
    
        #ici on verifie que nous n'avons pas deux fois le même outils dans le dico
        #si c'est le cas le booléen devient vrais
        #si c'est faux, une id est attribué à l'outil
        for val in d_id_tools.values() :

            if val == col.join(rep[i+1]) :
                exist = True

        if exist == False :
            d_id_tools[id] = col.join(rep[i+1])
            id += 1

    #on retourne deux dicos
    return d_id_tools, d_process_tools


#-----------------------------------------------------------------------------------------------------------------
#------------Traitement de la structure : on sépare sommets + identification et arrêtes + def partenaire----------
#-----------------------------------------------------------------------------------------------------------------
def Traitement_Structure(structure_workflow) :
    #on prend en entré le fichier structure workflow afin de traiter les sommets et les arrêtes
    #le fichier Json nous servira pour obtenir les dico id_tools et process_tools

    with open(structure_workflow, "r") as graph : 
        lines = graph.readlines()
    

    rep_process = repertoire_process("new_analysis_nf")
    

    y = len(lines)

    id_sommet = []
    arrete = []
    id = 0

    #---------------------- Traitement des lignes du fichier
    for line in lines[3:y-1] :
        line = line.rstrip("\n")
        line = line.strip("\t")
        
        z = len(line)
        id = 0
        i = True
        compteur = 0
        
        # on détermine quelle ligne est un sommet et 
        # quelle ligne est une arrête
        while compteur != z :
            for char in line[0:z] :
                compteur += 1
                if char == ">" :
                    i = False
        
        # Si True c'est un sommet, on le met donc dans un dico et
        # son identifiant correspondra à l'outils qu'il utilise
        if i == True :
            
            crochet = 0
            for char in line[0:z] : #ici on va épurer les lignes de sommet pour seulement garder le nom avant les crochets
                crochet += 1

                if char == "[" : 

                    for val in rep_process : 
                        #on va comparer les process obtenus dans le
                        #json avec les différentes lignes du fichier struture workflow
                        
                        if line[0:crochet-2] == val :
                            id_sommet.append([id,line[0:crochet-2]])
                        id+=1
        
        # Si False c'est une arrête, on le met dans une liste que l'on
        # traitera par la suite
        else :
            arrete.append(line)
    
    taille = 0
    partenaire = []

    for arr in range(0,len(arrete)) :

        if arrete[arr] != arrete[arr-1] :
            fin1 = 0
            fin2 = 0
            elle = arrete[arr]
        
        
            for char in elle :
                
                taille += 1
                    # boucle pour déterminer la fin du premier partenaire
                    # on s'arrête au 1er "-" qu'on croise (pour les fichiers de ce type)
                    # à partir de là on sait que la fin réel tu premiers partenaire de l'arrête
                    # est à 2 charactère du "-" et qu'il commence à 0.
                if char == ">" : 
                        
                    fin1 = taille-3

                    partenaire.append(elle[0:fin1])

                    # Même raisonnement que pour le premier partenaire, on détermine la fin en fonction
                    # du char "[" et on sait que la fin réelle est -2 charatère et le début à +2 de la fin
                    # estimé du premier
                elif char == "[" : ## boucle 
                       
                    fin2 = taille-2
                    partenaire.append(elle[fin1+4:fin2])
            taille = 0

    #on récupère une liste de liste avec les id des sommets et le sommet
    #et on récupère un fichier avec les duo des arrêtes (elles se suivent deux à
    #deux)
    
    
    return id_sommet, partenaire


#-----------------------------------------------------------------------------------------------------------------
#------------Table de correspondance process ==> ID --------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------

#Pour chaque process dans le dossier analysé, un identifiant lui sera assigné
#Si cela est une operation alors elle portera le même identifiant
def repertoire_process (dossier) :
    wb = op.Workbook()
    wb['Sheet'].title = 'id_process'

    ws1 = wb['id_process']
    ws1['A1'] = "ID"
    ws1['B1'] = 'name_process'

    row_start = 1
    col_start = 1
    id = 0
    id_process = list()
    dossier = t.triage(dossier)
    ##test sur un fichier

    for y in dossier:
        
        for path, subdirs, files in os.walk(y):
            for name in files:
                if name == "structure_worklow" : 
                    with open(os.path.join(path, name), "r") as data:
                        lines = data.readlines()           
                    y= len(lines)

                    for line in lines[3:y-1] :
                            line = line.rstrip("\n")
                            line = line.strip("\t")

                            z = len(line)
                                    
                            process = True
                            present = False
                            compteur = 0

                            while compteur != z :
                                for char in line[0:z] :
                                    compteur += 1
                                    if char == ">" :
                                        process = False
                                if process == True :
                                    crochet = 0
                                    for char in line[0:z] : 
                                            #ici on va épurer les lignes de sommet pour 
                                            #seulement garder le nom avant les crochets
                                        crochet += 1

                                        if char == "[" : 

                                            for i in id_process :
                                                if i == line[0:crochet-2] :
                                                    
                                                    present = True
                                            
                                            if present == False :
                                                
                                                id_process.append(line[0:crochet-2])
                                                ws1.cell(row_start+ id, col_start+1).value = line[0:crochet-2]
                                                ws1.cell(row_start+id, col_start).value = id
                                            
                                                
                                                
                                                id +=1

    wb.save('id_process.xlsx')
    return id_process



#-----------------------------------------------------------------------------------------------------------------
#---------------------NOUVEL FONCTION DE PARSING------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------
def new_new_Parsing(dossier) :

    import os
    import fonction_label as fl
    import pandas as pd
    import json

    no_tools = fl.dico_no_tools("new_new_analysis_nf") 
    # on extrait un dico contenant tous les process ne contenant pas d'outil

    label_ope = fl.dico_label_type_operation("new_new_analysis_nf")
    # on extrait un dico contenant le label de tous les types d'operation présent dans la base

    for path, subdirs, filenames in os.walk(dossier) :
        for name in filenames :
        
        # on va extraire les données à partir de certain fichier contenus dans le dossier de chaque workflow

            
            
        
            if name == "processes_info.json" :
                Tools = fl.Tool_Or_Not(os.path.join(path, name))
                
                # on extrait du fichier processes_info.json une liste de booléen ranger par ordre d'apparition de chaque sommet
                # True : le process a un au moins un outil / False : le process n'a pas d'outil
                
                
            if name == "structure_worklow" :

                nb_sommet = 0 #indicatif de l'id de chaque sommet

                with open(os.path.join(path, name) ,"r") as data_workflow :
                    lines = data_workflow.readlines()
            
                # le fichier sctructure_worklow contient les principales informations du workflow : ses sommets et ses arrêtes

                id_sommet = [] 
                #liste qui contiendra le nom du process, sa postion dans le workflow (ID) et son label (il correspond au groupe
                # de similarité d'outil dans lequel il est)

                arrete = []
                #liste qui contiendra les arrêtes (non traité dans cette première partie du code)

                l_lines = len(lines)
                
                

                for line in lines[3:l_lines-1] :
                    line = line.rstrip("\n")
                    line = line.strip("\t")

                    

                    l_line = len(line)
                    vertex_or_edge = True
                    character = 0
                    
                # Parmis toutes les lignes du fichier on va uniquement lire à partir de la ligne n=3 jusqu'à la ligne n-1 (avec n étant le nombre
                # total de ligne dans le fichier) + on nettoie les lignes pour enlever la tabulation et les retours à ligne présent dans les chaines
                # de caractère

                    while character != l_line :
                        for char in line[0:l_line] :
                            character += 1
                            if char == ">" :
                                vertex_or_edge = False
                    
                    # Ici on va différencier les sommets des arrêtes en cherchant dans chaque ligne le caractère ">" :
                    # si i = True ==> c'est un sommet
                    #si i = False ==> c'est une arrêtes

                    if vertex_or_edge == True :
                        line = line.split(" ")
                        sommet = line[0]
                        
                    # on continue de nettoyer la ligne afin de garder seulement le nom du sommet 

                        # ici on cherche à ne prendre que les process et non les OPERATIONs (le and est necessaire car il existe
                        # dans les données des OPERATIONs négatives et leur écriture n'est pas la même)

                        if sommet[0:9] != "OPERATION" and sommet[0:9] != '"OPERATIO' :

                            #on rentre dans la boucle ==> on a un process
                           
                            val = Tools[nb_sommet]
                            
                            

                            #on regarde si ce process a un outil


                            if val == True :

                                # val = TRUE , le process a au moins un outil, on va donc regarder dans quel groupe de similarité il se trouve
                                #dans la table de similarité.
                                #Pour ce faire, on a besoin du nom du process ainsi que le workflow dans lequel il se trouve d'ou l'intéret des
                                #lignes qui vont suivre.

                                sommet_lie = os.path.join(path, sommet)
                                sommet_lie=sommet_lie.split("__")
                                sommet_lie_un = sommet_lie[1].replace("/", ":")


                                # On va parcourir la table de similarité afin de trouver ou se trouve process et ainsi lui attribuer un label
                                data = pd.read_csv('groups_sim_nf_names.csv')
 
                                for i in range(0,len(data),1) :
                                
                                # i correspond au i-ième groupe de process == label

                                    rows = data.at[i, 'groups']
                                    rows = rows.split(",")
                                
                                    for process in rows  :
                                        
                                        pro = process.split("'")
                                        
                                        for proc in range(1,len(pro),2) : 


                                            if sommet_lie_un == pro[proc] :


                                                # si le sommet coorespond à un process de la table on va lui attribuer
                                                # le label du groupe
                                                # l'id est attribué par ordre d'arrivé

                                                id_sommet.append([sommet,nb_sommet,i])
                                                

                                nb_sommet +=1

                            else :
                                
                                # Si val = False alors on a un process qui n'utilise pas d'outil
                                # Dans ce cas la soit on attribue un label different à chaque process soit on leur
                                # attribue le même
                                
                                '''
                                ## Part for no _tools with different labels ###

                                # 

                                for label_no_tools in range(100000,100000+len(no_tools),1):
                                    if no_tools[label_no_tools] == sommet_lie_un :
                                        
                                        
                                        id_sommet.append([sommet,nb_sommet, label_no_tools])
                                    
                                ## -------------------------------------------------
                                '''

                                ## Part for no_tools with same labels ###

                                id_sommet.append([sommet,nb_sommet, 100000])
                                

                                ## -------------------------------------------

                                nb_sommet +=1
                                
                        
                        else :
                            name = "operations_extracted.json"
                            with open(os.path.join(path, name)) as data_operation :
                                file = json.load(data_operation)
                            
                                
                            action = file[sommet]["string"]
                            action = action.split(" ")
                            propre = action[0].split(".")
                            plus_que_propre = propre[0]
                            plus_que_propre= plus_que_propre.strip('\n')
                            

                            for lab_ope in range(10000, 10000+len(label_ope),1) :
                                if label_ope[lab_ope] == plus_que_propre :
                                    
                                    id_sommet.append([sommet, nb_sommet, lab_ope])
                                    

                            nb_sommet += 1
    ## a ce niveau on a tous les sommets avec un id et un label en fonction de leur catégories
                            
                    else :            
                        arrete.append(line)
        
    taille = 0
    partenaire = []

    for arr in range(0,len(arrete)) :

        if arrete[arr] != arrete[arr-1] :
            fin1 = 0
            fin2 = 0
            elle = arrete[arr]
            
            
            for char in elle :
                
                taille += 1
                        # boucle pour déterminer la fin du premier partenaire
                        # on s'arrête au 1er "-" qu'on croise (pour les fichiers de ce type)
                        # à partir de là on sait que la fin réel tu premiers partenaire de l'arrête
                        # est à 2 charactère du "-" et qu'il commence à 0.
                if char == ">" : 
                            
                    fin1 = taille-3

                    partenaire.append(elle[0:fin1])

                        # Même raisonnement que pour le premier partenaire, on détermine la fin en fonction
                        # du char "[" et on sait que la fin réelle est -2 charatère et le début à +2 de la fin
                        # estimé du premier
                elif char == "[" : ## boucle 
                            
                    fin2 = taille-2
                    partenaire.append(elle[fin1+4:fin2])
            taille = 0

    
    
    return id_sommet, partenaire

